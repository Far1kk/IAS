from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
import pickle
from bs4 import BeautifulSoup
import re
import time
from openpyxl import load_workbook
from random import uniform
from random_user_agent.user_agent import UserAgent
from random_user_agent.params import SoftwareName, OperatingSystem
from datetime import date as libdate
import pandas as pd


PATH_DRIVER = r"C:\Users\bogdan\parser_wordstat\driver\chromedriver.exe"
PATH_COOKIE = "cookies.pkl"
software = [SoftwareName.CHROME.value]
operating_sys = [OperatingSystem.WINDOWS.value]
user_agent_rotator = UserAgent(software_names=software, operating_system=operating_sys, limit=100)

url = "https://wordstat.yandex.ru/#!/history?words=БПЛА"
words_list = list()
with open("words.txt", 'r', encoding="utf-8") as f:
    for w in f.readlines():
        print(w)
        words_list.append(str(w).strip().replace("\n", ""))

class GetDriver:

    def __init__(self, path_driver, path_cookie):
        self.path_driver = path_driver
        self.path_cookie = path_cookie

    def driver(self, user_agent):
        try:
            service = ChromeService(executable_path=self.path_driver)
            options = webdriver.ChromeOptions()
            options.add_argument("start-maximized")
            options.add_argument(f"user-agent={user_agent}")
            options.add_argument(r"user-data-dir=C:\Users\bogdan\parser_wordstat\cock")
            driver = webdriver.Chrome(options=options, service=service)
            #driver.get(url=u"https://wordstat.yandex.ru/#!/history?words=бумага")
            """cookies = pickle.load(open(self.path_cookie, "rb"))
            for cookie in cookies:
                print(cookie)
                driver.add_cookie(cookie)"""
            return driver
        except Exception as ex:
            print(ex)
            driver.close()
            driver.quit()


class Scraping:
    """Парсинг истории, частот запросов с сайта WordStat."""

    def __init__(self, driver):
        self.driver = driver

    def freq(self, word, limit=1):
        """Принимает одно анализируемое слово, ставит его ключевым (key_word), выводит его частоту и частоту
        результирующих слов за прошедший месяц. Метод возвращает словарь, где первая пара- keyword:слово, следущие пары-
        слово:{текущая дата: частота}. Пример: {'key_word': 'word1', 'word1': {'30.12.23': int}...}
        limit- количество слов, включая ключевое, которые войдут в словарь."""
        try:
            daynow = libdate.today().strftime('%d.%m.%Y')
            phrase = list()
            freq = list()
            data = {'key_word': word}
            url = 'https://wordstat.yandex.ru/#!/?words=' + str(word).replace(' ', '%20') + '%2Bкупить'
            self.driver.get(url=url)
            time.sleep(uniform(2, 10))
            raw = self.driver.page_source
            soup = BeautifulSoup(raw, 'html.parser')
            phrase = soup.find_all('a', class_='b-link b-phrase-link__link', limit=limit)
            freq = soup.find_all('td', class_='b-word-statistics__td b-word-statistics__td-number', limit=limit)
            for o, e in zip(phrase, freq):
                data[o.get_text()] = {f'{daynow}': e.get_text().replace(' ', '')}
            print(data)
            return data
        except Exception as ex:
            print(ex)
            self.driver.close()
            self.driver.quit()


    def history(self, wordData):
        """Принимает словарь состоящий из ключевого слово и образованных от него результирующих слов Вордстата
        wordData = {'key_word': 'word1', 'word1': {'30.12.23': int}, 'word2': {'30.12.23': int}, 'word3': {'30.12.23': int}}
        На выходе словарь слова и истории data = {'key_word': 'word1', 'word1': {'30.10.23': int, '30.11.23': int,
        '30.12.23': int}, 'word2': {'30.10.23': int, '30.11.23': int, '30.12.23': int}} """
        try:
            for w in list(wordData.keys())[1:]:
                lastdata = wordData[w]
                url = 'https://wordstat.yandex.ru/#!/history?words=' + str(w).replace(' ', '%20')
                self.driver.get(url=url)
                time.sleep(uniform(2, 10))
                raw = self.driver.page_source
                soup = BeautifulSoup(raw, 'html.parser')
                odd = soup.find_all('tr', class_='odd')
                even = soup.find_all('tr', class_='even')
                predata = dict()
                for o, e in zip(odd, even):
                    match = re.sub('</span><span class="b-history__number-part">', '', str(o))
                    mas = re.findall(r"\d\d[0-9.,-]+", str(match))
                    predata[mas[1]] = mas[2]
                    match = re.sub('</span><span class="b-history__number-part">', '', str(e))
                    mas = re.findall(r"\d\d[0-9.,-]+", str(match))
                    predata[mas[1]] = mas[2]
                predata.update(lastdata)
                wordData[w] = predata
            print(wordData)
            return wordData
        except Exception as ex:
            print(ex)

    def uploadhistory(self, mas, path):
        fn = path
        wb = load_workbook(fn)
        ws = wb['sheet1']
        ws.append(mas)
        wb.save(fn)
        wb.close()
        print('Загрузка данных завершена')

    def uploadfreq(self, mas, path):
        fn = path
        wb = load_workbook(fn)
        ws = wb['sheet1']
        for w in mas.items():
            ws.append(list(w))
            print(w)
            wb.save(fn)
        wb.close()
        print('Загрузка данных завершена')

    def createdata(self, mas):
        zero = list()
        datelist = list()
        for i in mas:
            datelist.append(i)
        data = pd.DataFrame(zero, columns=datelist)
        data.to_excel('test.xlsx', sheet_name='sheet1', index=False)
        return datelist


if __name__ == '__main__':
    # odd, even = sraping(url)
    # masmerge = beauty(odd, even, num=1, object='object')
    # merg(masmerge)
    # for url in url_list:
    #    odd, even = sraping(url)
    #    masdata = beauty(odd, even, num=2, object='объект')
    #    uploadata(masdata)
    authdriver = GetDriver(PATH_DRIVER, PATH_COOKIE)
    authdriver = authdriver.driver(user_agent=user_agent_rotator.get_random_user_agent())
    k = 1
    mas = list()
    for i in words_list:
        print(i)
        data = Scraping(authdriver)
        wordata = data.freq(i, limit=1)
        mas.append(data.history(wordata))
        k += 1
        if k > 2:
            break
            authdriver.close()
            authdriver.quit()
            authdriver = GetDriver(PATH_DRIVER, PATH_COOKIE)
            authdriver = authdriver.driver(user_agent=user_agent_rotator.get_random_user_agent())
            k = 1
    """{'key_word': 'word1', 'word1': {'30.10.23': int, '30.11.23': int,
                                        '30.12.23': int}, 'word2': {'30.10.23': int, '30.11.23': int, '30.12.23': int}"""
    words = list()
    months = list()
    freqs = list()
    l = 0
    for k in range(0, len(mas)):
        for i in list(mas[k].keys())[1:]:
            words.append(i)
            l+=1
            if l <= 1:
                months = list(mas[0][i].keys())
    WORDSDATA = pd.DataFrame([], index=words, columns=months)
    print(words)
    print(months)
    num=0
    for x in WORDSDATA.index:
        for y in WORDSDATA.columns:
            print(mas[num][x][y])
            WORDSDATA.loc[x, y] = mas[num][x][y]
        num += 1

    with pd.ExcelWriter('WORDS.xlsx', engine='openpyxl', mode='w') as wrt:
        WORDSDATA.to_excel(wrt, sheet_name='History')

    """
    stat1 = list()
    stat2 = list()

    for i in list(mas[0].keys())[1:]:
        for n in mas[0][i].keys():
            stat1.append(int(mas[0][i][n]))
    for i in list(mas[1].keys())[1:]:
        for n in mas[1][i].keys():
            stat2.append(int(mas[1][i][n]))
    print(stat1)
    print(stat2)
    print(np.corrcoef(stat1, stat2)[0,1])
    print(acf(stat1))"""
    """{'key_word': 'word1', 'word1': {'30.10.23': int, '30.11.23': int,
                                    '30.12.23': int}, 'word2': {'30.10.23': int, '30.11.23': int, '30.12.23': int}"""
    # for i in range(0, len(words_list)-1, 5):
    #     print(i)
    #     words = list()
    #     words = words_list[:5]
    #     print(words)
    #     authdriver = GetDriver(PATH_DRIVER, PATH_COOKIE)
    #     authdriver = authdriver.driver(user_agent=user_agent_rotator.get_random_user_agent())
    #     print(user_agent_rotator.get_random_user_agent())
    #     data = Scraping(authdriver)
    #     data.freq(words)
    #     del words_list[0:5]

    # data = Scraping(authdriver)
    # data.history(words_list, "test.xlsx")
"""with open('datatext.txt', 'w') as f:
    for i in data:
        i = str(i)
        f.write(i)"""
